import openpyxl
from lxml import etree, builder
import imgkit
import pandas as pd
import pickle

class SpanningCell:
    def __init__(self, content, rows, columns):
        self.content = content
        self.row_range = range(rows[0], rows[1] + 1)
        self.col_range = range(columns[0], columns[1] + 1)
        self.origin = (rows[0], columns[0])
    
    def __str__(self):
        return f"Content: {self.content}, Rows: {self.row_range}, Columns: {self.col_range}"

    def contains(self, coordinate):
        return (coordinate[0] in self.row_range) and (coordinate[1] in self.col_range)

# Canonical data structure for the table physical structure
# typically the output of the detection + segmentation process
# can be converted from and to JSON, HTML, CSV, spreadsheet
class PhysicalStructure:
    def __init__(self, cells):
        # list of cells ordered ordered left-right, top-to-bottom
        self.cells = sorted(cells, key=lambda cell: (cell.origin[0], cell.origin[1]))
        self.shape = (self.cells[-1].row_range[-1], self.cells[-1].col_range[-1])

    @classmethod
    def from_azure(cls, cells_json):
        cells = []
        for cell in cells_json:
            r, c, rs, cs = cell['row_index'], cell['column_index'], cell['row_span'], cell['column_span']
            r, c = r + 1, c + 1 # 0-indexed to 1-indexed
            content = None if cell['content'] == "" else cell['content']
            cells.append(SpanningCell(content, [r, r+rs-1], [c, c+cs-1]))
        return cls(cells)

    @classmethod
    # every editted cell wil be considered, so only edit the area of the table
    def from_spreadsheet(cls, file_path):
        sheet = openpyxl.load_workbook(file_path).active
        cells = []
        # Process merged cells first
        for merged_range in sheet.merged_cells.ranges:
            r1, c1, r2, c2 = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            value = sheet.cell(row = r1, column = c1).value
            cells.append(SpanningCell(value, [r1-1, r2-1], [c1-1, c2-1]))
        done_cells = {(r, c) for cell in cells for r in cell.row_range for c in cell.col_range}
        # Iterate over all cells
        for row in sheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                r, c = cell.row-1, cell.column-1
                if (r, c) not in done_cells:
                    cells.append(SpanningCell(cell.value, [r, r], [c, c]))

        return cls(cells)

    def getCell(self, coordinate):
        for cell in self.cells:
            if cell.contains(coordinate):
                return cell
        else:
            return None
    
    def getRow(self, row, only_origin = False):
        if only_origin:
            row_cells = [cell for cell in self.cells if row == cell.row_range[0]]
        else:
            row_cells = [cell for cell in self.cells if row in cell.row_range]
        return sorted(row_cells, key=lambda cell: cell.origin[1]) if row_cells else []
    
    def to_HTML(self, prettify = False):
        table = builder.E.table()
        for row in range(1, self.shape[0] + 1):
            tr_elements = []
            for cell in self.getRow(row, only_origin=True):
                attributes = {}
                if len(cell.row_range) > 1:
                    attributes['rowspan'] = str(len(cell.row_range))
                if len(cell.col_range) > 1:
                    attributes['colspan'] = str(len(cell.col_range))
                content = "" if cell.content is None else str(cell.content)
                tr_elements.append(builder.E.td(content, **attributes))
            table.append(builder.E.tr(*tr_elements, i = str(row)))
        if prettify:
            table.set('style', 'border-collapse: collapse; width: 100%;')
            for element in table.iter('td', 'th'):
                element.set('style', 'border: 1px solid black; padding: 10px;')
        return etree.tostring(table, method='html', encoding='utf-8')

###################################################################################################################
###################################################################################################################
    
class Attribute:
    def __init__(self, name, origin, col_range):
        self.name = name
        self.origin = origin
        self.range = col_range
        self.parent = None
    
    def __str__(self):
        return f"Name: {self.name}, Origin: {self.origin}, Range: {self.range}, Parent: {self.parent.name if self.parent else None}"

    def setParent(self, parent):
        self.parent = parent

    def getParentChain(self):
        # Base case
        if not self.parent:
            return [self.name]
        # Recursive case
        return self.parent.getParentChain() + [self.name]

# Canonical data structure for the table logical structure
class LogicalStructure:
    def __init__(self, physical_structure, header_rows, entries_rows):
        # {attribute_index:Attribute, ...}
        self.base_attributes = {} 
        # {entry_index:{attribute_index:value,...},...} 
        # the value can be: a string, a 1D list, a 2D list or None
        self.entries = {} 
        self.physical_structure = physical_structure
        self.header_range = range(header_rows[0], header_rows[1] + 1)
        self.entries_range = {i:range(rows[0], rows[1] + 1) for i, rows in entries_rows.items()}

    def parse(self):
        self.parseHeader()
        self.parseEntries()

    # I am assuming the base attributes span is not reaching outside the header
    def parseHeader(self):
        # consider the whole bottom row as base attributes
        bottom_row_cells = self.physical_structure.getRow(self.header_range[-1])
        for i, cell in enumerate(bottom_row_cells):
            self.base_attributes[i+1] = Attribute(cell.content, cell.origin, cell.col_range)
        # map parent attributes
        pending_attributes = list(self.base_attributes.values())
        while pending_attributes:
            attribute = pending_attributes.pop(0)
            row, col = attribute.origin
            higher_cell = self.physical_structure.getCell((row-1, col))
            # cell exist, its origin is within the limits of the header, it has content, then set parent
            if higher_cell and higher_cell.origin[0] >= self.header_range[0] and higher_cell.content: 
                parent = Attribute(higher_cell.content, higher_cell.origin, higher_cell.col_range)
                attribute.setParent(parent)
                pending_attributes.append(parent)

    def parseEntries(self):
        for e_index, e_range in self.entries_range.items():
            self.entries[e_index] = {}
            for a_index, attribute in self.base_attributes.items():
                value = []
                covered_cells = []
                for row in e_range:
                    value.append([])
                    for column in attribute.range:
                        cell = self.physical_structure.getCell((row, column))
                        if cell not in covered_cells:
                            value[-1].append(cell.content)
                            covered_cells.append(cell)
                self.entries[e_index][a_index] = value
        # simplify
        for key, sub_dict in self.entries.items():
            for sub_key, value in sub_dict.items():
                # Remove empty lists
                value = [x for x in value if x]
                # Unnest length one lists
                if len(value) == 1:
                    if len(value[0]) == 1:
                        self.entries[key][sub_key] = value[0][0]
                    else:
                        self.entries[key][sub_key] = value[0]

    def to_dataframe(self):
        def sortedValues(dict): # guarantee the iteration is ordered - don't rely on dictionary insertion order 
            return [value for key, value in sorted(dict.items(), key = lambda item: item[0])]
        # Get parents for MultiIndex
        column_tuples = [attr.getParentChain() for attr in sortedValues(self.base_attributes)]
        # Rename None attributes to a synthetic string
        column_tuples = [["_unnamed_" if item is None else item for item in tup] for tup in column_tuples]
        # Normalize padding
        max_length = max(len(tup) for tup in column_tuples)
        column_tuples = [tup + [""] * (max_length - len(tup)) for tup in column_tuples]
        # Prepare Data for Rows
        data = [sortedValues(entry) for entry in sortedValues(self.entries)]
        return pd.DataFrame(data, columns = pd.MultiIndex.from_tuples(column_tuples), dtype = object)


if __name__ == "__main__":
    teste = 4
    if teste == 1:
        physical_structure = PhysicalStructure.from_spreadsheet('E:/Git/TableUnderstanding/example_files/header.xlsx')
        imgkit.from_string(physical_structure.to_HTML(prettify = True), "azure_test_dec.png")
    elif teste == 2:
        physical_structure = PhysicalStructure.from_spreadsheet('E:/Git/TableUnderstanding/example_files/header.xlsx')
        print(physical_structure.to_HTML())
    ################################################################################################
    elif teste == 3 or teste == 4:
        physical_structure = PhysicalStructure.from_spreadsheet('E:/Git/TableUnderstanding/example_files/header.xlsx')
        logical_structure = LogicalStructure()
        logical_structure.setHeaderRange([1, 3])
        logical_structure.parseHeader(physical_structure)
        if teste == 3:
            for i, attribute in logical_structure.base_attributes.items():
                print(i, attribute)
        else:
            logical_structure.setEntriesRange({
                1:(5, 5),
                2:(6, 6),
                3:(7, 7),
                4:(8, 8),
                5:(9, 9),
                6:(11, 11),
                7:(12, 12),
                8:(13, 13),
                9:(14, 15)
            })
            logical_structure.parseEntries(physical_structure)
            #print(logical_structure.entries)
            df = logical_structure.to_dataframe()
            print(df)
            with open('E:/Git/TableUnderstanding/example_files/df.pkl', 'wb') as f:
                pickle.dump(df, f)
    else:
        pass